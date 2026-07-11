import io
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dart_lib import account_map

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

AMOUNT_METRICS = account_map.METRIC_ORDER  # 9개 금액 항목
RATIO_METRICS = ["ROE(%)", "부채비율(%)"]
ALL_ROW_LABELS = AMOUNT_METRICS + RATIO_METRICS

AMOUNT_FORMAT = "#,##0"
RATIO_FORMAT = "0.00"


def _style_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER


def _write_value(cell, value, is_ratio):
    if value is None:
        cell.value = "N/A"
        cell.alignment = CENTER
        return
    cell.value = value
    cell.number_format = RATIO_FORMAT if is_ratio else AMOUNT_FORMAT


def _autofit(ws, ncols, width=16):
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def _write_company_sheet(wb, company, years, data_by_year):
    ws = wb.create_sheet(_safe_sheet_name(company))

    _style_header_cell(ws.cell(row=1, column=1, value="연도"))
    for j, label in enumerate(ALL_ROW_LABELS, start=2):
        _style_header_cell(ws.cell(row=1, column=j, value=label))

    for i, year in enumerate(years, start=2):
        ws.cell(row=i, column=1, value=f"{year}년").font = BOLD_FONT
        metrics = data_by_year.get(year, {})
        for j, label in enumerate(ALL_ROW_LABELS, start=2):
            is_ratio = label in RATIO_METRICS
            _write_value(ws.cell(row=i, column=j), metrics.get(label), is_ratio)

    ws.freeze_panes = "B2"
    _autofit(ws, len(ALL_ROW_LABELS) + 1)
    return ws


def _write_summary_sheet(wb, results, companies, years):
    ws = wb.create_sheet("비교요약", 0)

    _style_header_cell(ws.cell(row=1, column=1, value="항목"))
    _style_header_cell(ws.cell(row=2, column=1, value=""))
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for year in years:
        start_col = col
        for company in companies:
            _style_header_cell(ws.cell(row=2, column=col, value=company))
            col += 1
        end_col = col - 1
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        _style_header_cell(ws.cell(row=1, column=start_col, value=f"{year}년"))

    total_cols = col - 1

    for i, label in enumerate(ALL_ROW_LABELS, start=3):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        is_ratio = label in RATIO_METRICS
        col = 2
        for year in years:
            for company in companies:
                value = results.get(company, {}).get(year, {}).get(label)
                _write_value(ws.cell(row=i, column=col), value, is_ratio)
                col += 1

    ws.freeze_panes = "B3"
    _autofit(ws, total_cols)
    return ws


def build_workbook(results: dict, companies: list[str], years: list[str]) -> Workbook:
    """results: {회사명: {연도: {지표명: 값}}} 구조를 받아 Workbook 객체를 생성한다."""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    _write_summary_sheet(wb, results, companies, years)
    for company in companies:
        _write_company_sheet(wb, company, years, results.get(company, {}))

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_workbook(results: dict, companies: list[str], years: list[str], output_path: str) -> str:
    """results: {회사명: {연도: {지표명: 값}}} 구조를 받아 엑셀 파일로 저장한다."""
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb = build_workbook(results, companies, years)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # 더미 데이터로 서식/구성 확인
    companies = ["회사A", "회사B", "회사C"]
    years = ["2023", "2024", "2025"]
    dummy_results = {}
    for ci, company in enumerate(companies):
        dummy_results[company] = {}
        for yi, year in enumerate(years):
            metrics = {label: 1_000_000_000 * (ci + 1) * (yi + 1) for label in AMOUNT_METRICS}
            metrics["ROE(%)"] = 5.0 + ci + yi
            metrics["부채비율(%)"] = 50.0 + ci * 10
            dummy_results[company][year] = metrics
    dummy_results["회사B"]["2024"]["매출액"] = None  # N/A 케이스 확인

    path = write_workbook(dummy_results, companies, years, "output/_test_dummy.xlsx")
    print("생성됨:", path)
